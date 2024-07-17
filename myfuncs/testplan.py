from openpyxl import load_workbook, Workbook
import attrs
from myclasses import IVSweep
import pandas as pd

@attrs.define
class chnVoltBias:
    description     : str
    resource        : str
    V_force         : str
    I_compl         : str = attrs.field(
        default=1e-3,
        kw_only=True,
    )
    VoltSense       : bool = attrs.field(
        default=False,
        kw_only=True,
    )
    V_compl         : float = attrs.field(
        default=3,
        kw_only=True,
    )
    I_force         : float = attrs.field(
        default=0,
        kw_only=True,
    )




@attrs.define
class CurrentLimit:
    low_level = attrs.field(
        default=10e-6,
        kw_only=True,
    )
    high_level = attrs.field(
        default=1e-3,
        kw_only=True,
    )

@attrs.define
class CurrentLimitRange:
    low_level = attrs.field(
        default=100e-9,
        kw_only=True,
    )
    high_level = attrs.field(
        default=30e-3,
        kw_only=True,
    )


@attrs.define
class Resource:
    device: str
    GATE : int
    DRAIN: int
    SOURCE: int
    BULK: int

@attrs.define
class VtlinSweep:
    resource: Resource
    V_force_D: float
    V_force_S: float
    V_force_B: float
    I_compl_G: float
    I_compl_D: float
    I_compl_S: float
    I_compl_B: float
    V_step_G: float
    V_start_G: float
    V_stop_G: float
    Ith:    float
    width: float
    length: float
    # CurrentLimit: CurrentLimit
    # CurrentLimitRange: CurrentLimitRange
    sourceDelay: float = attrs.field(
        default=1e-3,
        kw_only=True,
    )
    apertureTime: float = attrs.field(
        default=1e-3,
        kw_only=True,
    )
    test_case: IVSweep = attrs.field(
        default=None,
        kw_only=True,
    )




@attrs.define
class VtSatSweep:
    resource: Resource
    V_force_D: float
    V_force_S: float
    V_force_B: float
    I_compl_G: float
    I_compl_D: float
    I_compl_S: float
    I_compl_B: float
    V_step_G: float
    V_start_G: float
    V_stop_G: float
    Ith : float
    width: float
    length: float
    CurrentLimit: CurrentLimit               =  attrs.field(
        default=CurrentLimit(),
        kw_only=True,
    )

    CurrentLimitRange: CurrentLimitRange     = attrs.field(
        default=CurrentLimitRange(),
        kw_only=True,
    )

    test_case: IVSweep = attrs.field(
        default=None,
        kw_only=True,
    )

@attrs.define
class WAT:
    resource: Resource
    description: str
    Vtlin: VtlinSweep
    Vtsat: VtSatSweep

    tile_coord: list = attrs.field(
        default=None,
        kw_only=True,
    )

    tileName: str = attrs.field(
        default='A1',
        kw_only=True,
    )
    port_map: dict = attrs.field(
        default=None,
        kw_only=True,
    )

def dumpExceltoPython(testplan, sitesNum):
    wb = load_workbook(filename=testplan)
    ws = wb.active
    dDut, dWAT, dTest_Vtlin, dTest_VtSat, dTile, dResource = {}, {}, {}, {}, {}, {}
    header = []
    rows_buf = []
    count = 0
    for row in ws.iter_rows(min_row=None, max_col=None, max_row=None, values_only=True):
        if count == 0:
            header = row
            # print(header)
        else:
            rows_buf.append(row)
        count += 1
    df_plan = pd.DataFrame(rows_buf, columns=header)
    norepeat_df_plan = df_plan.drop_duplicates(subset=['dut.name']).reset_index()
    # count_1 = 0
    sDuts = set()

    # Get all the DUTs and resources
    count = 0
    for dut in norepeat_df_plan.loc[:, 'dut.name']:
        # print(norepeat_df_plan.loc[:, ' dut.pads[G]'])
        padGATE = norepeat_df_plan.loc[count, 'dut.pads[G]']
        # print(f'Gate is {padGATE}')
        padDRAIN = norepeat_df_plan.loc[count, 'dut.pads[D]']
        padSOURCE = norepeat_df_plan.loc[count, 'dut.pads[S]']
        padBULK = norepeat_df_plan.loc[count, 'dut.pads[B]']
        tile_x = norepeat_df_plan.loc[count, 'tile.x0']
        tile_y = norepeat_df_plan.loc[count, 'tile.y0']
        '''
        we define the terminal sequence for dut is G,D,S,B
        so in terms of the matrix map, we do things like Gpad->SMU1, Dpad->SMU2, Spad->SMU3, Bpad->SMU4
        '''
        dResource[dut] = Resource(dut, padGATE, padDRAIN, padSOURCE, padBULK) # we define the terminal sequence for dut is G,D,S,B
        dTile[dut] = [tile_x, tile_y]
        sDuts.add(dut)
        count += 1
    # print(dResource)

    count = 0
    for c in df_plan.loc[:, 'fullname']:
        dut = df_plan.loc[count, 'dut.name']
        if 'VtLin' in c:
            vForceD = df_plan.loc[count,  'test.V_force_D']
            vForceS = df_plan.loc[count,  'test.V_force_S']
            vForceB = df_plan.loc[count,  'test.V_force_B']
            vStartG = df_plan.loc[count,  'test.V_start_G']
            vStopG = df_plan.loc[count,  'test.V_stop_G']
            vStepG = df_plan.loc[count,  'test.V_step_G'] 
            iComplG = df_plan.loc[count,  'test.I_compl_G']
            iComplD = df_plan.loc[count,  'test.I_compl_D']
            iComplS = df_plan.loc[count,  'test.I_compl_S']
            iComplB = df_plan.loc[count,  'test.I_compl_B']
            Ith = df_plan.loc[count,  'test.I_thresh_D']
            Wg = df_plan.loc[count,  'dut.Wg']
            Lg = df_plan.loc[count,  'dut.Lg']
            dTest_Vtlin[dut] = VtlinSweep(dResource[dut], vForceD, vForceS, vForceB, iComplG, iComplD, 
                                          iComplS, iComplB, vStepG, vStartG, vStopG, Ith=Ith, width=Wg, length=Lg,)
        elif 'VtSat' in c:
            vForceD = df_plan.loc[count,  'test.V_force_D']
            vForceS = df_plan.loc[count,  'test.V_force_S']
            vForceB = df_plan.loc[count,  'test.V_force_B']
            vStartG = df_plan.loc[count,  'test.V_start_G']
            vStopG = df_plan.loc[count,  'test.V_stop_G']
            vStepG = df_plan.loc[count,  'test.V_step_G'] 
            iComplG = df_plan.loc[count,  'test.I_compl_G']
            iComplD = df_plan.loc[count,  'test.I_compl_D']
            iComplS = df_plan.loc[count,  'test.I_compl_S']
            iComplB = df_plan.loc[count,  'test.I_compl_B']
            Ith = df_plan.loc[count,  'test.I_thresh_D']
            Wg = df_plan.loc[count,  'dut.Wg']
            Lg = df_plan.loc[count,  'dut.Lg']
            dTest_VtSat[dut] = VtSatSweep(dResource[dut], vForceD, vForceS, vForceB, iComplG, 
                                          iComplD, iComplS, iComplB, vStepG, vStartG, vStopG, Ith=Ith, width=Wg, length=Lg)
        
        count += 1

    # print("hahha")
    for dut in sDuts:
        resource = dResource[dut]
        Vtlin = dTest_Vtlin[dut]
        VtSat = dTest_VtSat[dut]
        # Ibmax = dTest_IbMAX[dut]
        dWAT[dut] = WAT(dResource[dut], dut, dTest_Vtlin[dut], dTest_VtSat[dut],tile_coord=dTile[dut])
    return dWAT
import attrs
import nidcpower
from openpyxl import load_workbook, Workbook
import pandas as pd
from typing import Dict
# from config import wiring
import numpy as np
import hightime
import time
import csv
# -*- coding: utf-8 -*-
# This file was generated
import array  # noqa: F401
# Used by @ivi_synchronized
from functools import wraps
from myfuncs.testplan import dumpExceltoPython, CurrentLimitRange, CurrentLimit, WAT
from myfuncs.myclasses import ChnVoltBias, ChnVoltSweep, IVSweep
import nidcpower._attributes as _attributes
import nidcpower._converters as _converters
import nidcpower._library_interpreter as _library_interpreter
import nidcpower.enums as enums
import nidcpower.errors as errors

import nidcpower.lcr_load_compensation_spot as lcr_load_compensation_spot  # noqa: F401

import nidcpower.lcr_measurement as lcr_measurement  # noqa: F401

import hightime
import collections
# from change_range_softly import *
from binary_search import runIVSweeps_softwareAutoRange
from myfuncs.current_range import find_range
from driver.P8XL import P8XL
from driver.E5250A import E5250A_Simple
from config import prober, initCoord

class WaferAcceptanceTest:
    

    def __init__(self, testplan, sitesNum):
        self.dWAT: Dict[str, WAT]
        self.dWAT = dumpExceltoPython(testplan, sitesNum=sitesNum)
        self.sw_define()
        self.current_limit = CurrentLimit()
        self.current_limit_range = CurrentLimitRange()
        self.sess = None
        # self.stressInterval = generateInterval(10000)
        
    def populate():
        wb = Workbook()    
        ws = wb.active
        ws['A1'] = 42
        ws.append([1, 2, 3])
        import datetime
        ws['A2'] = datetime.datetime.now()
        wb.save('sample.xlsx')

    
    # would it be fine if we dump all channels of 4163 into a self.session and we then open the relays we don't need?
    def common_settings(self):
        self.sess = nidcpower.Session(resource_name='SMU1', reset=True,  options = {'simulate': True, 'driver_setup': {'Model': '4163', 'BoardType': 'PXIe', }, })
        # print(self.sess.channels['SMU1/0,SMU1/1'])
        self.sess.voltage_level_autorange = True
        self.sess.current_limit_autorange = True
        self.sess.output_function = nidcpower.OutputFunction.DC_VOLTAGE
        self.sess.voltage_level = 0
        self.sess.current_limit = 1e-3
        self.sess.source_mode = nidcpower.SourceMode.SEQUENCE
        self.sess.output_connected = False # turn off all channels 
        self.sess.output_enabled = False
        self.sess.autorange = True
        self.sess.autorange_aperture_time_mode = nidcpower.AutorangeApertureTimeMode.AUTO
        # self.sess.sequence_step_delta_time_enabled = True # not supported in 4163
        
        return self.sess

    def makeupIVrequest_vtlin(self):
        for dut in self.dWAT:
            vStart, vStop, vStep = self.dWAT[dut].Vtlin.V_start_G, self.dWAT[dut].Vtlin.V_stop_G, self.dWAT[dut].Vtlin.V_step_G
            I_compl_G=self.dWAT[dut].Vtlin.I_compl_G
            
            I_compl_S=self.dWAT[dut].Vtlin.I_compl_S
            I_compl_B=self.dWAT[dut].Vtlin.I_compl_B
            w = self.dWAT[dut].Vtlin.width
            l = self.dWAT[dut].Vtlin.length
            Ith = self.dWAT[dut].Vtlin.Ith * w / l
            vForceD = self.dWAT[dut].Vtlin.V_force_D
            vForceS = self.dWAT[dut].Vtlin.V_force_S
            vForceB = self.dWAT[dut].Vtlin.V_force_B
            I_compl_D=find_range(Ith) #fix range
            vtlin = IVSweep(ChnVoltSweep('G', 'SMU1/0', V_start=vStart, V_stop=vStop, V_step=vStep, I_compl=I_compl_G, remote_sense=False),
                        [ChnVoltBias('D', 'SMU2/0', vForceD, I_compl=I_compl_D, remote_sense=False),
                        ChnVoltBias('S', 'SMU3/0', vForceS, I_compl=I_compl_S, VoltSense=False),
                        ChnVoltBias('B', 'SMU4/0', vForceB, I_compl=I_compl_B),
                        ],
                        apertureTime=20e-3,
                        sourceDelay=20e-3,
                        isMaster=1,
                        )
            # self.dWAT: Dict[str, WAT]
            
            self.dWAT[dut].Vtlin.test_case = vtlin
            # runIVSweeps_softwareAutoRange(vtlin) 


    def makeupIVrequest_vtsat(self):
        for dut in self.dWAT:
            vStart, vStop, vStep = self.dWAT[dut].Vtsat.V_start_G, self.dWAT[dut].Vtsat.V_stop_G, self.dWAT[dut].Vtsat.V_step_G
            I_compl_G=self.dWAT[dut].Vtsat.I_compl_G
            
            I_compl_S=self.dWAT[dut].Vtsat.I_compl_S
            I_compl_B=self.dWAT[dut].Vtsat.I_compl_B
            w = self.dWAT[dut].Vtsat.width
            l = self.dWAT[dut].Vtsat.length
            Ith = self.dWAT[dut].Vtsat.Ith * w / l
            vForceD = self.dWAT[dut].Vtsat.V_force_D
            vForceS = self.dWAT[dut].Vtsat.V_force_S
            vForceB = self.dWAT[dut].Vtsat.V_force_B
            I_compl_D=find_range(Ith) #fix range
            Vtsat = IVSweep(ChnVoltSweep('G', 'SMU1/0', V_start=vStart, V_stop=vStop, V_step=vStep, I_compl=I_compl_G, remote_sense=False),
                        [ChnVoltBias('D', 'SMU2/0', vForceD, I_compl=I_compl_D, remote_sense=False),
                        ChnVoltBias('S', 'SMU3/0', vForceS, I_compl=I_compl_S, VoltSense=False),
                        ChnVoltBias('B', 'SMU4/0', vForceB, I_compl=I_compl_B),
                        ],
                        apertureTime=20e-3,
                        sourceDelay=20e-3,
                        isMaster=1,
                        )
            self.dWAT[dut].Vtsat.test_case = Vtsat
            # runIVSweeps_softwareAutoRange(vtsat)


    def sw_define(self, sw:E5250A_Simple=None):
        for dut in self.dWAT:
            portMap = {
                E5250A_Simple.InputPort.SMU1: self.dWAT[dut].resource.GATE,
                E5250A_Simple.InputPort.SMU2: self.dWAT[dut].resource.DRAIN,
                E5250A_Simple.InputPort.SMU3: self.dWAT[dut].resource.SOURCE,
                E5250A_Simple.InputPort.SMU4: self.dWAT[dut].resource.BULK,
            }
            self.dWAT[dut].port_map = portMap
            # sw.setupPortMap(portMap)

    
    def sw_setup(self, dut, sw:E5250A_Simple):
        
        sw.setupPortMap(self.dWAT[dut].port_map)
        
    
    def run(self, prober_station:P8XL, dieX, dieY):
        
        curX, curY = initCoord
        for dut in self.dWAT:
            dx = self.dWAT[dut].tile_coord[0]-curX
            dy = self.dWAT[dut].tile_coord[1]-curY
            # self.common_settings()
            #check needle position and move prober
            with prober_station.connect():
                prober_station.getWaferParams()
                
                prober_station.driveDistanceX(dx)
                prober_station.driveDistanceY(dy)
                for i in range(2):
                    prober_station.downZ()
                    prober_station.upZ()
            self.sw_setup(dut, E5250A_Simple())
            curX, curY = self.dWAT[dut].tile_coord

            #setup sw port map

            vt_vtlin = runIVSweeps_softwareAutoRange(self.dWAT[dut].Vtlin.test_case)
            vt_vtSat = runIVSweeps_softwareAutoRange(self.dWAT[dut].Vtsat.test_case)
            record = {"X": dieX, "Y": dieY, "DUT": dut, "vt_vtlin": vt_vtlin, "vt_vtSat": vt_vtSat}
            with open('record.csv', 'a', newline='') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=record.keys())
                writer.writerow(record)


    

    




if __name__ == "__main__":
    from dies import dies
    prober_address = prober[1]
    p8 = P8XL(prober_address)
    wat = WaferAcceptanceTest('NSVT_NHVT.xlsx', 1)

    # for k, v in wat.dWAT.items():
    #     print(k, v)
    #     print("\n")
    portmap = wat.sw_setup()
    for k, v in portmap.items():
        print(k, v)

    for die in dies:
        
        wat.makeupIVrequest_vtlin()
        wat.makeupIVrequest_vtsat()
        with p8.connect():
            p8.getWaferParams()
            p8.moveToDie(die[0], die[1])
        wat.run(prober_address, die[0], die[1])

